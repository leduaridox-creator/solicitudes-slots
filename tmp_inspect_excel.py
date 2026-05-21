import os
import sys
try:
    import openpyxl
except Exception as e:
    print('MISSING_OPENPYXL', e)
    sys.exit(1)
files=[f for f in os.listdir('Formato Excel') if f.endswith('.xlsx')]
print('files=', files)
for f in files:
    path=os.path.join('Formato Excel', f)
    wb=openpyxl.load_workbook(path, data_only=True)
    ws=wb.active
    print('---', f, 'sheets:', wb.sheetnames, 'max_row', ws.max_row, 'max_col', ws.max_column)
    print('merged:', list(ws.merged_cells.ranges)[:20])
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), min_col=1, max_col=min(20, ws.max_column), values_only=True), 1):
        print(i, row)
